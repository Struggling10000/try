import os
import turtle as t
from string import digits
from openpyxl import load_workbook
import cv2
def floor(s):
    remove_digits = str.maketrans('', '', digits)
    res = s.translate(remove_digits)
    max_s_count = 0
    for i in res:
        if res.count(i) > max_s_count:
            max_s_count = res.count(i)
            if res.count(i) == max_s_count:
                max_s_count = res.count(i)
    return max_s_count
def excel_name(filename):
    wb = load_workbook(filename)
    ws = wb["Sheet1"]
    all_info = []
    column=ws['A']
    for cell in column:
        all_info.append(cell.value)


    for i in all_info[1:]:


        print(i)
        glynum=0
        for str in i:
            if str.isdigit():
                glynum=glynum+1
        if glynum>5:
            # i = i.replace('E5', '')
            # E_count=i.count('E')
            # if  E_count<5:
            test(i)
            # else: continue
        else: continue
def type_1(size=20):#绿圆形
    t.pendown()
    t.begin_fill()
    t.fillcolor((0,1,0))
    t.circle(size/2)
    t.end_fill()
    t.penup()
    t.left(90)
    t.fd(size)
    t.right(90)
def type_2(size=20):#正方形
    t.pendown()
    t.begin_fill()
    t.fillcolor(0, 0, 1)
    t.fd(size/2)
    i = 1
    while i <= 3:
        t.left(90)
        t.forward(size)
        i = i + 1
    t.left(90)
    t.fd(size/2)
    t.end_fill()
    t.penup()
    t.left(90)
    t.fd(size)
    t.right(90)
def type_3(size=20):#紫色菱形
    t.pendown()
    t.begin_fill()
    t.fillcolor(1, 0, 1)
    t.left(45)
    t.fd(size/2 *2**0.5)
    i=1
    for i in range(3):
        t.left(90)
        t.fd(size/2 * 2 ** 0.5)
    t.left(135)
    t.end_fill()
    t.penup()
    t.fd(size)
    t.right(90)
def type_4(size=20):#白色菱形
    t.pendown()
    t.left(45)
    t.fd(size/2  * 2 ** 0.5)
    i = 1
    for i in range(3):
        t.left(90)
        t.fd(size/2 * 2 ** 0.5)
    t.left(135)
    t.penup()
    t.fd(size)
    t.right(90)
def type_5(size=15):#红色三角
    t.pendown()
    t.begin_fill()
    t.fillcolor(1, 0, 0)
    t.left(30)
    for i in range(3):
        t.fd(size)
        t.right(120)
    t.right(30)
    t.end_fill()
    t.penup()
def type_6(size=20):  # 黄圆形
        t.pendown()
        t.begin_fill()
        t.fillcolor((1, 1, 0))
        t.circle(size / 2)
        t.end_fill()
        t.penup()
        t.left(90)
        t.fd(size)
        t.right(90)
def begin_draw():
    t.setup(1000, 1000, 200, 200)
    t.pensize(1)
    t.penup()
    t.speed(10000)
    t.hideturtle()

def end_draw(str):
    # save pic [, font=("font-name", font_size, "font_type")
    # t.hideturtle()
    #保存图片

    ts = t.getscreen()
    FileName =str+ ".eps"
    ts.getcanvas().postscript(file=FileName)
    t.write(str)
    # t.done()    ##### 将画面固定在屏幕
    t.reset()  #####   清除画布
    os.popen('cd desktop\My_code2 && magick -density 300 {F} {E}.jpg'.format(F=FileName, E=str))
def delete_five(str):
    str = str.replace('B5b', '')
    str = str.replace('C5c', '')
    str = str.replace('D2', '')
    str = str.replace('D2d', '')
    str = str.replace('E5e', '')
    str = str.replace('A2', '')
    str = str.replace('B2', '')
    str = str.replace('C1', '')
    str = str.replace('a', '')
    str = str.replace('b', '')
    str = str.replace('c', '')
    str = str.replace('d', '')
    str = str[2:]
    str_list = str.split('D1')
    str_list.reverse()
    return str_list
def defult_pain(str):
      # 后续再设定不同阈值，对五糖核心分支的角度长度进行特异性设定。
    point = []
    t.goto(0,-75)
    type_2()
    t.pendown()
    t.left(90)
    t.fd(30)
    t.right(90)
    type_2()
    t.pendown()
    t.left(90)
    t.fd(30)
    t.right(90)
    type_1()
    tloc = t.pos()
    t.pendown()
    floor_count = floor(delete_five(str)[0]) + floor(delete_five(str)[1])
    if floor_count < 7:
        #最多一层少于7个时
        t.fd(5)
        t.left(30)            #应修改
        t.fd(80)              #应修改
        t.right(30)
        type_1()
        tloc_right=t.pos()
        t.goto(tloc)
        t.pendown()
        t.left(180)
        t.fd(5)
        t.right(30)            #应修改
        t.fd(80)               #应修改
        t.right(150) #应修改
        t.penup()
        type_1()
        tloc_left = t.pos()
    if floor_count >= 7:
        # 最多一层多余7个时
        t.fd(5)
        t.left(30)  # 应修改
        t.fd(100)  # 应修改
        t.right(30)
        type_1()
        tloc_right = t.pos()
        t.goto(tloc)
        t.pendown()
        t.left(180)# 应修改
        t.fd(5)
        t.right(30)
        t.fd(100)  # 应修改
        t.right(150)  # 应修改
        t.penup()
        type_1()
        tloc_left=t.pos()
    point.append(tloc_right)
    point.append(tloc_left)
    return point
def find_f(str):
    sta=str.count('B5b')
    sta2=str.count('C5c')
    if sta==3:
            t.goto(10,-65)
            t.pendown()
            t.fd(10)
            t.penup()
            type_5()
            t.goto(-10, -65)
            t.left(180)
            t.pendown()
            t.fd(10)
            t.penup()
            type_5()
            t.goto(0, -75)
            t.left(90)
            t.pendown()
            t.fd(10)
            t.penup()
            type_5()
            t.left(90)
    if sta==2:
            t.goto(10,-65)
            t.pendown()
            t.fd(10)
            t.penup()
            type_5()
            t.goto(-10,-65)
            t.left(180)
            t.pendown()
            t.fd(10)
            t.penup()
            type_5()
            t.right(180)
    if  sta==1:
            t.goto(10,-65)
            t.pendown()
            t.fd(10)
            t.penup()
            type_5()

    if sta2==2:
            t.goto(10,-15)
            t.pendown()
            t.fd(10)
            t.penup()
            type_5()
            t.goto(-10, -15)
            t.left(180)
            t.pendown()
            t.fd(10)
            t.penup()
            type_5()
            t.right(180)
    if  sta2==1:
            t.goto(10,-15)
            t.pendown()
            t.fd(10)
            t.penup()
            type_5()
def find_bisecto(str):
    if 'D2d' in str:
        t.goto(0,45)
        t.left(90)
        t.pendown()
        t.fd(40)
        t.penup()
        t.right(90)
        type_2()
    if 'D2E5' in str:
        t.goto(0,45)
        t.left(90)
        t.pendown()
        t.fd(40)
        t.penup()
        t.right(90)
        type_2()
        t.pendown()
        t.fd(10)
        t.right(90)
        t.fd(10)
        t.left(90)
        t.fd(20)
        t.penup()
        type_5()
        # t.goto()
    if 'D2E5eE5' in str:
        t.goto(0,45)
        t.left(90)
        t.pendown()
        t.fd(40)
        t.penup()
        t.right(90)
        type_2()
        t.pendown()
        t.fd(10)
        t.right(90)
        t.fd(10)
        t.left(90)
        t.fd(20)
        t.penup()
        type_5()
        t.left(180)
        t.fd(40)
        t.pendown()
        t.fd(20)
        t.penup()
        type_5()
        t.right(180)
    if 'D2E5eE5eE5' in str:
        t.goto(0,45)
        t.left(90)
        t.pendown()
        t.fd(40)
        t.penup()
        t.right(90)
        type_2()
        t.pendown()
        t.fd(10)
        t.right(90)
        t.fd(10)
        t.left(90)
        t.fd(20)
        t.penup()
        type_5()
        t.left(180)
        t.fd(40)
        t.pendown()
        t.fd(20)
        t.penup()
        type_5()
        t.right(180)
        t.fd(30)
        t.left(90)
        t.fd(10)
        t.pendown()
        t.fd(20)
        t.penup()
        type_5()
        t.right(90)
def draw_type1(c):
    #将绿园==1，高甘露
    for case in c:
        if case=='1':
            type_1()
            break
        if case=='2':
            type_2()
            break
        if case=='3':
            type_3()
            break
        if case=='4':
            type_4()
            break
        if case=='5':
            type_5()
            break
        if case=='6':
            type_3()
            break
        if case=='':  # default
            print('no type!')
def draw_type2(c):
    #将黄圆==1，非高甘露
    for case in c:
        if case=='1':
            type_6()
            break
        if case=='2':
            type_2()
            break
        if case=='3':
            type_3()
            break
        if case=='4':
            type_4()
            break
        if case=='5':
            type_5()
            break
        if case=='6':
            type_3()
            break
        if case=='7':
            type_4()
            break
        if case=='':  # default
            print('no type!')
# def painting_brance(b):
#
#     #直线
#     if b==1:
#         t.pendown()
#         t.fd(30)
#         t.right(30)
#         t.penup()
#         tloc2=t.pos()
#         return tloc2
#
#     #分叉
#     elif b==2:
#         tloc=t.pos()
#         t.pendown()
#         t.fd(5)
#         t.left(30)
#         t.fd(45)
#         t.penup()
#         tloc2=t.pos()
#         t.goto(tloc)
#
#         t.pendown()
#         t.left(150)
#         t.fd(5)
#         t.right(30)
#         t.fd(45)
#         t.right(150)
#         t.penup()
#
#         return tloc2
def find_loc(str,c):    #是用来 确定同一级（两个分支）例如 D1###dD2###d这样的形式中两个D的位置 ， 结果为  [0,18] 这样的形式
    count = 0
    strloc=[]
    for echar in str:
        count=count+1
        if echar == c:
            strloc.append(count-1)
    return strloc

def p_brance(str):
    # 若要将左侧三分支 向左倾斜，请添加[1]
    if len(str)>0:
        loc = find_loc(str, str[0])
        if len(loc) == 1:
            t.left(90)
            t.pendown()
            t.fd(30)
            t.right(90)
            t.penup()
            tloc2 = t.pos()
            return tloc2
        if len(loc) == 2:
            tloc = t.pos()
            t.pendown()
            t.fd(5)
            t.left(35)        #应修改
            t.fd(55)#应修改
            t.penup()
            tloc2 = t.pos()
            t.goto(tloc)
            t.pendown()
            t.left(145)#应修改
            t.fd(5)
            t.right(35)
            t.fd(55)#应修改
            t.right(145)#应修改
            t.penup()
            return tloc2
        if len(loc)==3 and floor(delete_five(str)[0])>=4:
            a=[]
            tloc = t.pos()
            t.pendown()
            t.fd(5)
            t.left(20)        #应修改
            t.fd(90)#应修改
            t.left(10)
            t.penup()
            tloc2 = t.pos()
            t.goto(tloc)
            t.left(60)
            t.pendown()
            t.fd(30)
            t.penup()
            tloc3= t.pos()
            t.goto(tloc)
            t.left(90)
            t.pendown()
            t.fd(5)
            t.right(25)
            t.fd(70)

            t.penup()
            t.right(155)
            a.append(tloc2)
            a.append(tloc3)
            return a
        if len(loc) == 3  and floor(delete_five(str)[0])<4:
            a = []
            tloc = t.pos()
            t.pendown()
            t.fd(5)
            t.left(33)  # 应修改
            t.fd(55)  # 应修改
            t.penup()
            tloc2 = t.pos()
            t.goto(tloc)
            t.left(57)
            t.pendown()
            t.fd(30)
            t.penup()
            tloc3 = t.pos()
            t.goto(tloc)
            t.left(90)
            t.pendown()
            t.fd(5)
            t.right(33)
            t.fd(55)
            t.penup()
            t.right(147)
            a.append(tloc2)
            a.append(tloc3)
            return a
        if len(loc)==3 and floor(delete_five(str)[1])>=4:
            a=[]
            tloc = t.pos()
            t.pendown()
            t.fd(5)
            t.left(20)        #应修改
            t.fd(90)#应修改
            t.left(10)
            t.penup()
            tloc2 = t.pos()
            t.goto(tloc)
            t.left(60)
            t.pendown()
            t.fd(30)
            t.penup()
            tloc3= t.pos()
            t.goto(tloc)
            t.left(90)
            t.pendown()
            t.fd(5)
            t.right(25)
            t.fd(70)

            t.penup()
            t.right(155)
            a.append(tloc2)
            a.append(tloc3)
            return a
        if len(loc) == 3  and floor(delete_five(str)[1])<4:
            a = []
            tloc = t.pos()
            t.pendown()
            t.fd(5)
            t.left(33)  # 应修改
            t.fd(55)  # 应修改
            t.penup()
            tloc2 = t.pos()
            t.goto(tloc)
            t.left(57)
            t.pendown()
            t.fd(30)
            t.penup()
            tloc3 = t.pos()
            t.goto(tloc)
            t.left(90)
            t.pendown()
            t.fd(5)
            t.right(33)
            t.fd(55)
            t.penup()
            t.right(147)
            a.append(tloc2)
            a.append(tloc3)
            return a

def anay_structofgly_(ogly_str,GP_list,turtleloc):
    if len(ogly_str)==0 and len(GP_list)==0:
        return 0
    elif len(ogly_str) == 3 and len(GP_list) == 0: # 最后一个的话，直接绘制
        # t.pendown()
        # t.left(90)
        # t.fd(30)
        # t.right(90)
        # t.penup()
        draw_type2(ogly_str[1])
        return 0
    elif len(ogly_str) == 3 and len(GP_list) != 0:
        # t.pendown()
        # t.left(90)
        # t.fd(30)
        # t.right(90)
        # t.penup()
        draw_type2(ogly_str[1])
        t.goto(turtleloc.pop())
        temp_str = GP_list.pop()
        return anay_structofgly_(temp_str, GP_list, turtleloc)
    elif len(ogly_str)==0 and len(GP_list)!=0:
        t.goto(turtleloc.pop())
        temp_str = GP_list.pop()
        return anay_structofgly_(temp_str, GP_list, turtleloc)
    elif len(ogly_str)!=0:
        loc = find_loc(ogly_str, ogly_str[0])
        if len(loc)==1:
            ogly_str1 = ogly_str
            draw_type2(ogly_str1[1])
            tloc3=t.pos()
        elif len(loc)==2:
            ogly_str2 = ogly_str[loc[1]:]
            ogly_str1 = ogly_str[0:loc[1]]
            GP_list.append(ogly_str2)
            if len(ogly_str1) == 3:
                return anay_structofgly_(ogly_str1, GP_list, turtleloc)
            draw_type2(ogly_str1[1])
            tloc3 = t.pos()
        elif len(loc)==3:
            ogly_str2 = ogly_str[loc[1]:loc[2]] #将 第二个分支 切出来放入 栈中
            ogly_str3 = ogly_str[loc[2]:]      #将 第三个分支 切出来放入 栈中
            ogly_str1 = ogly_str[0:loc[1]]
            GP_list.append(ogly_str3)
            GP_list.append(ogly_str2)
            if len(ogly_str1) == 3:
                return anay_structofgly_(ogly_str1, GP_list, turtleloc)
            draw_type2(ogly_str1[1])
            tloc3 = t.pos()

        '''
        检查是否有岩藻糖
        '''

        Stang = ogly_str1[2] + '5' + ogly_str1[-2].lower()
        sta=ogly_str1.count(Stang)
        # if sta==0:
        #     ogly_str1 = ogly_str1[2:-1]
        #     return anay_structofgly_(ogly_str1, list, turtleloc)
        if sta==1:
            t.goto(tloc3)
            t.fd(10)
            t.right(90)
            t.fd(10)
            t.left(90)
            t.pendown()
            t.fd(12)
            t.penup()
            type_5()
            t.goto(tloc3)
        elif sta==2:
            t.goto(tloc3)
            t.pendown()
            t.fd(10)
            t.right(90)
            t.fd(10)
            t.left(90)
            t.fd(13)
            t.penup()
            type_5()
            t.left(180)
            t.fd(33)
            t.pendown()
            t.fd(13)
            t.pendown()
            type_5()
            t.right(180)
            t.goto(tloc3)
        elif sta==3:
            t.goto(tloc3)
            t.pendown()
            t.fd(10)
            t.right(90)
            t.fd(10)
            t.left(90)
            t.fd(10)
            t.penup()
            type_5()
            t.left(180)
            t.fd(30)
            t.pendown()
            t.fd(10)
            t.pendown()
            type_5()
            t.right(90)
            t.goto(tloc3)
            t.pendown()
            t.fd(10)
            t.penup()
            type_5()
            t.right(90)
        Stang = ogly_str1[2] + '5' + ogly_str1[-2].lower()
        ogly_str_noF = ogly_str1.replace(Stang, '')
        if Stang in ogly_str1:
            if len(ogly_str_noF)==3 and GP_list==[]:
                return 0
            if len(ogly_str_noF)==3 and GP_list!=[]:
                ogly_str1 = ogly_str_noF[2:-1]
                return anay_structofgly_(ogly_str1, GP_list, turtleloc)
            loc2 = find_loc(ogly_str_noF, ogly_str_noF[2])
        elif Stang not in ogly_str1:
                loc2 = find_loc(ogly_str_noF, ogly_str_noF[2])
        if len(loc2)==1:
            t.left(90)
            t.pendown()
            t.fd(30)
            t.right(90)
            t.penup()
            # if Stang in ogly_str1:
            #     ogly_str1 = ogly_str_noF[2:-1]
            #     return anay_structofgly_(ogly_str1, GP_list, turtleloc)
            # else:
            #     ogly_str1 = ogly_str1[2:-1]
            #     return anay_structofgly_(ogly_str1, GP_list, turtleloc)
        elif len(loc2)==2:
            tloc = t.pos()
            t.pendown()
            t.fd(4)
            t.left(60)          #应修改
            t.fd(35)            #应修改
            t.penup()
            tloc2 = t.pos()
            t.goto(tloc)
            t.pendown()
            t.left(120)
            t.fd(4)
            t.right(60)      #应修改
            t.fd(35)        #应修改
            t.right(120)    #应修改
            t.penup()
            turtleloc.append(tloc2)
        if Stang in ogly_str1:
            ogly_str1 = ogly_str_noF[2:-1]
            return anay_structofgly_(ogly_str1, GP_list, turtleloc)
        else:
            ogly_str1 = ogly_str1[2:-1]
            return anay_structofgly_(ogly_str1, GP_list, turtleloc)
def anay_structofgly_man(ogly_str,GP_list,turtleloc):
    if len(ogly_str)==0 and len(GP_list)==0:  # 这个list是 出现分支时，将后一个分支切片，放入list，画完左分支糖链后 进行list中有分支糖链的绘画。
        return 0
    elif len(ogly_str) == 3 and len(GP_list) == 0:  # 最后一个的话，直接绘制
        draw_type1(ogly_str[1])
        return 0
    elif len(ogly_str) == 3 and len(GP_list) != 0:
        draw_type1(ogly_str[1])
        to=turtleloc.pop()
        t.goto(to)
        temp_str = GP_list.pop()
        return anay_structofgly_man(temp_str, GP_list, turtleloc)
    elif len(ogly_str)==0 and len(GP_list)!=0:
        t.goto(turtleloc.pop())
        temp_str = GP_list.pop()
        return anay_structofgly_man(temp_str, GP_list, turtleloc)
    elif len(ogly_str)!=0:
        loc = find_loc(ogly_str, ogly_str[0])
        if len(loc)==1:
            ogly_str1 = ogly_str
            draw_type1(ogly_str1[1])
            # tloc3=t.pos()
        elif len(loc)==2:
            ogly_str2 = ogly_str[loc[1]:] #将 第二个分支 切出来放入 栈中
            ogly_str1 = ogly_str[0:loc[1]]
            # tloc3=t.pos()
            GP_list.append(ogly_str2)
            if len(ogly_str1) == 3:
                return anay_structofgly_man(ogly_str1, GP_list, turtleloc)
            draw_type1(ogly_str1[1])
        elif len(loc)==3:
            ogly_str2 = ogly_str[loc[1]:loc[2]] #将 第二个分支 切出来放入 栈中
            ogly_str3 = ogly_str[loc[2]:]      #将 第三个分支 切出来放入 栈中
            ogly_str1 = ogly_str[0:loc[1]]
            GP_list.append(ogly_str2)
            GP_list.append(ogly_str3)
            if len(ogly_str1) == 3:
                return anay_structofgly_man(ogly_str1, GP_list, turtleloc)
            draw_type1(ogly_str1[1])

        loc2 = find_loc(ogly_str1, ogly_str1[2])
        if len(loc2)==1:
            t.left(90)
            t.pendown()
            t.fd(30)
            t.right(90)
            t.penup()
            ogly_str1 = ogly_str1[2:-1]
            return anay_structofgly_man(ogly_str1, GP_list, turtleloc)
        elif len(loc2)==2:
            tloc = t.pos()
            t.pendown()
            t.fd(4)
            t.left(60)
            t.fd(35)
            t.penup()
            tloc2 = t.pos()
            t.goto(tloc)
            t.pendown()
            t.left(120)
            t.fd(4)
            t.right(60)
            t.fd(35)
            t.right(120)
            t.penup()
            turtleloc.append(tloc2)
            ogly_str1 = ogly_str1[2:-1]
            return anay_structofgly_man(ogly_str1, GP_list, turtleloc)
        #
        #     '''
        #     检查是否有岩藻糖
        #     '''
        # Stang = ogly_str1[2] + '5' + ogly_str1[-2].lower()
        # sta=ogly_str1.count(Stang)
        # if sta==0:
        #     ogly_str1 = ogly_str1[2:-1]
        #     return anay_structofgly_man(ogly_str1, list, turtleloc)
        # elif sta==1:
        #     t.goto(tloc3)
        #     t.fd(10)
        #     t.right(90)
        #     t.fd(10)
        #     t.left(90)
        #     t.pendown()
        #     t.fd(20)
        #     t.penup()
        #     type_5()
        #     t.goto(tloc)
        #     index = ogly_str1.index(Stang)
        #     ogly_str1 = ogly_str1[0:index] + ogly_str1[index + 3:-1] + ogly_str1[-1]  # 删除岩藻糖
        #     ogly_str1 = ogly_str1[2:-1]
        #     return anay_structofgly_man(ogly_str1, list, turtleloc)
        # elif sta==2:
        #     t.goto(tloc3)
        #     t.pendown()
        #     t.fd(10)
        #     t.right(90)
        #     t.fd(10)
        #     t.left(90)
        #     t.fd(30)
        #     t.penup()
        #     type_5()
        #     t.left(180)
        #     t.fd(50)
        #     t.pendown()
        #     t.fd(30)
        #     t.pendown()
        #     type_5()
        #     t.right(180)
        #     t.goto(tloc)
        #     index = ogly_str1.index(Stang)
        #     ogly_str1 = ogly_str1[0:index] + ogly_str1[index + 3:-1] + ogly_str1[-1]  # 删除岩藻糖
        #     ogly_str1 = ogly_str1[2:-1]
        #     return anay_structofgly_man(ogly_str1, list, turtleloc)
        # elif sta==3:
        #     t.goto(tloc3)
        #     t.pendown()
        #     t.fd(10)
        #     t.right(90)
        #     t.fd(10)
        #     t.left(90)
        #     t.fd(30)
        #     t.penup()
        #     type_5()
        #     t.left(180)
        #     t.fd(50)
        #     t.pendown()
        #     t.fd(30)
        #     t.pendown()
        #     type_5()
        #     t.right(90)
        #     t.goto(tloc)
        #     t.pendown()
        #     t.fd(30)
        #     t.penup()
        #     index = ogly_str1.index(Stang)
        #     ogly_str1 = ogly_str1[0:index] + ogly_str1[index + 3:-1] + ogly_str1[-1]  # 删除岩藻糖
        #     ogly_str1 = ogly_str1[2:-1]
        #     return anay_structofgly_man(ogly_str1, list, turtleloc)

def test(str):
    P=delete_five(str)
    if P[0].count('E')>3 or P[1].count('E')>3 :
        print('一侧超出了三个分支')
    else:
        GP_list = []
        turtleloc = []
        begin_draw()
        #绘画五糖核心
        a = defult_pain(str)
        find_f(str)
        find_bisecto(str)
        #绘画除五糖核心，平分行，核心岩藻糖之外的 结构。
        t.goto(a[1])
        p_left=(p_brance(delete_five(str)[1]))
        if type(p_left) is not list:        ###############################
            turtleloc.append(p_left)
        if type(p_left) is list:
            turtleloc.append(p_left[1])
            turtleloc.append(p_left[0])
        if '2'  not in delete_five(str)[1]:
            anay_structofgly_man(delete_five(str)[1], GP_list, turtleloc)
        else:
            anay_structofgly_(delete_five(str)[1], GP_list, turtleloc)

        t.goto(a[0])
        p_right=(p_brance(delete_five(str)[0]))
        if type(p_right) is not list:        ###############################
            turtleloc.append(p_right)
        if type(p_right) is list:
            turtleloc.append(p_right[0])
            turtleloc.append(p_right[1])
        if '2' not in delete_five(str)[0]:
            anay_structofgly_man(delete_five(str)[0],GP_list, turtleloc)
        else:
            anay_structofgly_(delete_five(str)[0], GP_list, turtleloc)
        end_draw(str)



if __name__ == "__main__":
    # N_str6 = 'A2B2C1D1E2F1G5gfF5fF5fedD1E2F1feE2F1G5gfF5fedcbB5bB5ba'
    # filename='糖型.xlsx'
    # excel_name(filename)
    # test('A2B2C1D1E1F1fF1feE1F1fF1feE1F1fF1fedD1E1eE2F2fF2feE2F2fF2feE2F2fF2fedcba')
    # test('A2B2C1D1E2edD1E2F1feE2F1G3gfF3feE2F1G3gfF3fedcba')
    test('A2B2C1D1E2F1G6g3fedD1E2F1fF3fedcba')











